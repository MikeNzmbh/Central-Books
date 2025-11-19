import csv
import io

from django.contrib import admin
from django.http import HttpResponse
from django.urls import reverse
from django.utils import timezone
from django.utils.http import urlencode

try:
    from openpyxl import Workbook
except ImportError:  # pragma: no cover - optional dependency for XLSX export
    Workbook = None

from .models import (
    Account,
    BankAccount,
    Business,
    Category,
    Customer,
    Expense,
    Invoice,
    Supplier,
    Item,
    JournalEntry,
)


DANGEROUS_PREFIXES = ("=", "+", "-", "@")


@admin.register(Business)
class BusinessAdmin(admin.ModelAdmin):
    # === List display / UX ===
    list_display = ("id", "name", "currency", "owner_user", "created_at")
    list_display_links = ("id", "name")
    readonly_fields = ("id", "created_at")
    ordering = ("-id",)
    list_filter = ("currency", "created_at")
    search_fields = ("name", "owner_user__username", "currency")
    date_hierarchy = "created_at"
    list_per_page = 25

    # === Actions ===
    actions = ["export_selected_as_csv"]

    @admin.action(description="Export selected businesses to CSV")
    def export_selected_as_csv(self, request, queryset):
        return self._export_queryset_as_csv(queryset)

    # === “Export all (filtered)” buttons ===
    change_list_template = "admin/core/business/change_list.html"

    def changelist_view(self, request, extra_context=None):
        """
        If &_export=csv/xlsx is present, export ALL rows that match the current
        filters/search/sort. We must remove `_export` from request.GET before
        constructing the ChangeList, otherwise Django tries to filter by it.
        """
        export_fmt = request.GET.get("_export")

        if export_fmt in {"csv", "xlsx"}:
            original_get = request.GET
            cleaned_get = request.GET.copy()
            cleaned_get.pop("_export", None)
            request.GET = cleaned_get
            try:
                changelist = self.get_changelist_instance(request)
                queryset = changelist.get_queryset(request)
            finally:
                request.GET = original_get

            return (
                self._export_queryset_as_csv(queryset)
                if export_fmt == "csv"
                else self._export_queryset_as_xlsx(queryset)
            )

        base = reverse("admin:core_business_changelist")
        query = request.GET.copy()
        query.pop("_export", None)
        csv_url = f"{base}?{query.urlencode()}&_export=csv"
        xlsx_url = f"{base}?{query.urlencode()}&_export=xlsx"

        extra = {
            "export_all_csv_url": csv_url,
            "export_all_xlsx_url": xlsx_url,
        }
        extra_context = (extra_context or {}) | extra
        return super().changelist_view(request, extra_context=extra_context)

    # === Helpers ===
    def _export_queryset_as_csv(self, queryset):
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="businesses.csv"'
        writer = csv.writer(response)
        writer.writerow(
            [
                "ID",
                "Name",
                "Currency",
                "Owner Username",
                "Owner Email",
                "Created At",
            ]
        )
        for business in queryset:
            writer.writerow(
                [
                    business.id,
                    business.name,
                    business.currency,
                    getattr(business.owner_user, "username", ""),
                    getattr(business.owner_user, "email", ""),
                    timezone.localtime(business.created_at).strftime("%Y-%m-%d %H:%M:%S"),
                ]
            )
        return response

    def _export_queryset_as_xlsx(self, queryset):
        workbook = Workbook() # type: ignore
        worksheet = workbook.create_sheet("Businesses")
        if workbook.sheetnames.count("Sheet") > 0:
            workbook.remove(workbook["Sheet"])
        worksheet.append(
            [
                "ID",
                "Name",
                "Currency",
                "Owner Username",
                "Owner Email",
                "Created At",
            ]
        )
        for business in queryset:
            worksheet.append(
                [
                    business.id,
                    business.name,
                    business.currency,
                    getattr(business.owner_user, "username", ""),
                    getattr(business.owner_user, "email", ""),
                    timezone.localtime(business.created_at).strftime("%Y-%m-%d %H:%M:%S"),
                ]
            )

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="businesses.xlsx"'
        return response


@admin.register(BankAccount)
class BankAccountAdmin(admin.ModelAdmin):
    list_display = (
        "name",
        "business",
        "bank_name",
        "usage_role",
        "account",
        "account_number_mask",
        "is_active",
    )
    list_filter = ("usage_role", "is_active", "bank_name")
    search_fields = ("name", "bank_name", "account_number_mask", "business__name")
    autocomplete_fields = ("business", "account")


@admin.register(Account)
class AccountAdmin(admin.ModelAdmin):
    list_display = ("code", "name", "business", "type", "is_active")
    search_fields = ("code", "name", "business__name")
    list_filter = ("type", "is_active")
    autocomplete_fields = ("business", "parent")
